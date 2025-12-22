"""
금시세 데이터 통계 분석 및 엑셀 기록
"""

import pandas as pd
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# 엑셀 파일 읽기
filename = 'gold_price_1year_20251222.xlsx'
print(f"'{filename}' 파일을 읽는 중...")

df = pd.read_excel(filename)

print(f"\n✓ {len(df)}개의 데이터를 불러왔습니다.")
print(f"\n컬럼 목록:")
for col in df.columns:
    print(f"  - {col}")

# 통계 계산할 주요 컬럼들
stats_columns = ['매입가_순금(3.75g)', '매도가_순금(3.75g)', '매도가_18K', '매도가_14K']

# 통계 데이터 생성
stats_data = []

for col in stats_columns:
    if col in df.columns:
        stats_data.append({
            '항목': col,
            '평균': round(df[col].mean(), 2),
            '최대값': df[col].max(),
            '최소값': df[col].min(),
            '중앙값': round(df[col].median(), 2),
            '표준편차': round(df[col].std(), 2)
        })

stats_df = pd.DataFrame(stats_data)

print("\n통계 요약:")
print(stats_df.to_string(index=False))

# 날짜 정보
if '고시날짜' in df.columns:
    print(f"\n데이터 기간:")
    print(f"  시작: {df['고시날짜'].iloc[-1]}")
    print(f"  종료: {df['고시날짜'].iloc[0]}")
    print(f"  총 일수: 약 {len(df)}건")

# 새 엑셀 파일로 저장 (원본 + 통계)
output_filename = 'gold_price_1year_with_stats_20251222.xlsx'

with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
    # 원본 데이터 시트
    df.to_excel(writer, sheet_name='금시세 데이터', index=False)
    
    # 통계 시트
    stats_df.to_excel(writer, sheet_name='통계 요약', index=False)
    
    # 워크북 가져오기
    workbook = writer.book
    
    # 통계 시트 스타일링
    stats_sheet = workbook['통계 요약']
    
    # 헤더 스타일
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for cell in stats_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 컬럼 너비 조정
    stats_sheet.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        stats_sheet.column_dimensions[col].width = 15
    
    # 숫자 포맷 (천단위 구분)
    for row in stats_sheet.iter_rows(min_row=2, max_row=len(stats_data)+1, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = '#,##0.00'
            cell.alignment = Alignment(horizontal='right')

print(f"\n✓ 통계가 포함된 파일이 '{output_filename}'로 저장되었습니다.")
print(f"\n파일 구성:")
print(f"  - 시트1: '금시세 데이터' ({len(df)}개 행)")
print(f"  - 시트2: '통계 요약' (평균, 최대값, 최소값, 중앙값, 표준편차)")
